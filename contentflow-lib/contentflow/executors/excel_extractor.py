"""Excel extraction executor using openpyxl for sheets, tables, and images."""

import base64
import io
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel extraction. "
        "Install it with: pip install openpyxl"
    )

from . import ParallelExecutor
from ..models import Content

logger = logging.getLogger("contentflow.executors.excel_extractor")


class ExcelExtractorExecutor(ParallelExecutor):
    """
    Extract content from Excel workbooks using openpyxl.
    
    This executor analyzes Excel workbooks (.xlsx, .xlsm) to extract text, tables,
    sheet-level data, cell values, and embedded images using the openpyxl library.
    
    Configuration (settings dict):
        - extract_text (bool): Extract full text content from all cells
          Default: True
        - extract_sheets (bool): Create sheet-level data chunks
          Default: True
        - extract_tables (bool): Extract tables from sheets (defined tables or data ranges)
          Default: True
        - extract_properties (bool): Extract workbook properties
          Default: False
        - extract_images (bool): Extract embedded images from sheets
          Default: False
        - extract_formulas (bool): Extract cell formulas in addition to values
          Default: False
        - extract_comments (bool): Extract cell comments
          Default: False
        - content_field (str): Field containing Excel bytes
          Default: "content"
        - temp_file_path_field (str): Field containing temp file path
          Default: "temp_file_path"
        - output_field (str): Field name for extracted data
          Default: "excel_output"
        - image_output_mode (str): How to store extracted images
          Default: "base64"
          Options: "base64", "bytes"
        - sheet_separator (str): Separator between sheets in full text
          Default: "\n\n===\n\n"
        - include_empty_cells (bool): Include empty cells in extraction
          Default: False
        - max_rows_per_table (int): Maximum rows to extract per table (0 = unlimited)
          Default: 0
        - max_columns_per_table (int): Maximum columns to extract per table (0 = unlimited)
          Default: 0
        - read_only (bool): Open workbook in read-only mode for better performance
          Default: True
        - data_only (bool): Read cell values instead of formulas
          Default: True

        Also setting from ParallelExecutor and BaseExecutor apply.
    
    Example:
        ```python
        executor = ExcelExtractorExecutor(
            id="excel_extractor",
            settings={
                "extract_text": True,
                "extract_sheets": True,
                "extract_tables": True,
                "extract_images": True,
            }
        )
        ```
    
    Input:
        Document or List[Document] each with (ContentIdentifier) id containing:
        - data['content']: Excel workbook bytes, OR
        - data['temp_file_path']: Path to Excel file
        
    Output:
        Document or List[Document] with added fields:
        - data['excel_output']['text']: Full extracted text from all cells
        - data['excel_output']['sheets']: List of sheet data with cells and metadata
        - data['excel_output']['tables']: List of extracted tables (if enabled)
        - data['excel_output']['properties']: Workbook properties (if enabled)
        - data['excel_output']['images']: List of extracted images (if enabled)
        - data['excel_output']['formulas']: List of cell formulas (if enabled)
        - data['excel_output']['comments']: List of cell comments (if enabled)
        - summary_data['sheets_processed']: Number of sheets processed
        - summary_data['tables_extracted']: Number of tables extracted
    """
    
    def __init__(
        self,
        id: str,
        settings: Optional[Dict[str, Any]] = None,
        **kwargs
    ):
        super().__init__(
            id=id,
            settings=settings,
            **kwargs
        )
        
        # Extract configuration
        self.extract_text = self.get_setting("extract_text", default=True)
        self.extract_sheets = self.get_setting("extract_sheets", default=True)
        self.extract_tables = self.get_setting("extract_tables", default=True)
        self.extract_properties = self.get_setting("extract_properties", default=False)
        self.extract_images = self.get_setting("extract_images", default=False)
        self.extract_formulas = self.get_setting("extract_formulas", default=False)
        self.extract_comments = self.get_setting("extract_comments", default=False)
        self.content_field = self.get_setting("content_field", default="content")
        self.temp_file_field = self.get_setting("temp_file_path_field", default="temp_file_path")
        self.output_field = self.get_setting("output_field", default="excel_output")
        self.image_output_mode = self.get_setting("image_output_mode", default="base64")
        self.sheet_separator = self.get_setting("sheet_separator", default="\n\n===\n\n")
        self.include_empty_cells = self.get_setting("include_empty_cells", default=False)
        self.max_rows_per_table = self.get_setting("max_rows_per_table", default=0)
        self.max_columns_per_table = self.get_setting("max_columns_per_table", default=0)
        self.read_only = self.get_setting("read_only", default=True)
        self.data_only = self.get_setting("data_only", default=True)
        
        # Validate image output mode
        if self.image_output_mode not in ["base64", "bytes"]:
            raise ValueError(f"{self.id}: Invalid image_output_mode: {self.image_output_mode}. Must be 'base64' or 'bytes'")
        
        if self.debug_mode:
            logger.debug(
                f"ExcelExtractorExecutor {self.id} initialized: "
                f"text={self.extract_text}, sheets={self.extract_sheets}, "
                f"tables={self.extract_tables}, images={self.extract_images}"
            )
    
    async def process_content_item(
        self,
        content: Content
    ) -> Content:
        """Process a single Excel workbook using openpyxl.
           Implements the abstract method from ParallelExecutor.
        """
        
        try:
            if not content or not content.data:
                raise ValueError("Content must have data")
            
            # Get Excel content
            excel_bytes = None
            excel_path = None
            
            # Try to get from content field
            if self.content_field in content.data:
                excel_bytes = content.data[self.content_field]
            
            # Try to get from temp file
            elif self.temp_file_field in content.data:
                excel_path = content.data[self.temp_file_field]
            
            if not excel_bytes and not excel_path:
                raise ValueError(
                    f"Excel missing required content. "
                    f"Needs either '{self.content_field}' or '{self.temp_file_field}'"
                )
            
            if self.debug_mode:
                source = f"file: {excel_path}" if excel_path else f"bytes: {len(excel_bytes)} bytes"
                logger.debug(f"Processing Excel {content.id} from {source}")
            
            wb: Workbook = None
            
            # Open Excel workbook
            try:
                if excel_bytes:
                    wb = load_workbook(
                        io.BytesIO(excel_bytes),
                        read_only=self.read_only,
                        data_only=self.data_only
                    )
                else:
                    wb = load_workbook(
                        excel_path,
                        read_only=self.read_only,
                        data_only=self.data_only
                    )
            except InvalidFileException as e:
                logger.warning(
                    f"Invalid Excel file for content {content.id}: {str(e)}"
                )
                content.summary_data['excel_extraction_status'] = "invalid_file"
                return content
            
            extracted_data = {}
            
            # Extract sheets and tables
            all_text = []
            sheets_data = []
            all_tables = []
            all_formulas = []
            all_comments = []
            all_images = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Extract sheet data
                sheet_text, sheet_info = self._extract_sheet_data(sheet, sheet_name)
                
                if self.extract_text and sheet_text:
                    all_text.append(f"Sheet: {sheet_name}\n{sheet_text}")
                
                if self.extract_sheets:
                    sheets_data.append(sheet_info)
                
                # Extract tables from sheet
                if self.extract_tables:
                    sheet_tables = self._extract_tables_from_sheet(sheet, sheet_name)
                    all_tables.extend(sheet_tables)
                
                # Extract formulas
                if self.extract_formulas and not self.data_only:
                    sheet_formulas = self._extract_formulas_from_sheet(sheet, sheet_name)
                    all_formulas.extend(sheet_formulas)
                
                # Extract comments
                if self.extract_comments:
                    sheet_comments = self._extract_comments_from_sheet(sheet, sheet_name)
                    all_comments.extend(sheet_comments)
                
                # Extract images
                if self.extract_images:
                    sheet_images = self._extract_images_from_sheet(sheet, sheet_name, len(all_images))
                    all_images.extend(sheet_images)
            
            if self.extract_text:
                extracted_data['text'] = self.sheet_separator.join(all_text)
                if self.debug_mode:
                    logger.debug(f"Extracted {len(extracted_data['text'])} characters of text")
            
            if self.extract_sheets:
                extracted_data['sheets'] = sheets_data
                if self.debug_mode:
                    logger.debug(f"Processed {len(sheets_data)} sheets")
            
            if self.extract_tables:
                extracted_data['tables'] = all_tables
                if self.debug_mode:
                    logger.debug(f"Extracted {len(all_tables)} tables")
            
            if self.extract_formulas:
                extracted_data['formulas'] = all_formulas
                if self.debug_mode:
                    logger.debug(f"Extracted {len(all_formulas)} formulas")
            
            if self.extract_comments:
                extracted_data['comments'] = all_comments
                if self.debug_mode:
                    logger.debug(f"Extracted {len(all_comments)} comments")
            
            if self.extract_images:
                extracted_data['images'] = all_images
                if self.debug_mode:
                    logger.debug(f"Extracted {len(all_images)} images")
            
            # Extract workbook properties
            if self.extract_properties:
                properties = self._extract_workbook_properties(wb)
                extracted_data['properties'] = properties
                if self.debug_mode:
                    logger.debug(f"Extracted workbook properties")
            
            # Store extracted data
            content.data[self.output_field] = extracted_data
            
            # Update summary
            content.summary_data['sheets_processed'] = len(sheets_data) if self.extract_sheets else len(wb.sheetnames)
            content.summary_data['tables_extracted'] = len(all_tables)
            content.summary_data['images_extracted'] = len(all_images)
            content.summary_data['extraction_status'] = "success"
            
            # Close workbook
            wb.close()
                            
        except Exception as e:
            logger.error(
                f"ExcelExtractorExecutor {self.id} failed processing workbook {content.id}",
                exc_info=True
            )
            
            # Raise the exception to be handled upstream if needed
            raise
        
        return content
    
    def _extract_sheet_data(self, sheet: Worksheet, sheet_name: str) -> tuple[str, Dict[str, Any]]:
        """Extract data from a worksheet.
        
        Args:
            sheet: Worksheet to extract data from
            sheet_name: Name of the worksheet
            
        Returns:
            Tuple of (sheet_text, sheet_info_dict)
        """
        sheet_text_parts = []
        cell_data = []
        
        # Get the dimensions of the sheet
        if sheet.max_row is None or sheet.max_column is None:
            max_row = 0
            max_col = 0
        else:
            max_row = sheet.max_row
            max_col = sheet.max_column
        
        # Extract cell values
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_values = []
            for cell in row:
                value = cell.value
                if value is not None:
                    row_values.append(str(value))
                    if self.include_empty_cells or value:
                        cell_data.append({
                            "cell": cell.coordinate,
                            "value": value,
                            "row": cell.row,
                            "column": cell.column
                        })
                elif self.include_empty_cells:
                    row_values.append("")
            
            if row_values:
                sheet_text_parts.append("\t".join(row_values))
        
        sheet_text = "\n".join(sheet_text_parts)
        
        sheet_info = {
            "sheet_name": sheet_name,
            "max_row": max_row,
            "max_column": max_col,
            "cell_count": len(cell_data),
            "char_count": len(sheet_text),
        }
        
        # Only include cell data if it's reasonable size
        if len(cell_data) <= 1000:  # Limit to prevent huge output
            sheet_info["cells"] = cell_data
        else:
            sheet_info["cells_truncated"] = True
            sheet_info["cells"] = cell_data[:1000]
        
        return sheet_text, sheet_info
    
    def _extract_tables_from_sheet(self, sheet: Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract tables from a worksheet.
        
        Args:
            sheet: Worksheet to extract tables from
            sheet_name: Name of the worksheet
            
        Returns:
            List of table dictionaries
        """
        tables = []
        
        # Extract defined tables
        if hasattr(sheet, 'tables') and sheet.tables:
            for table_name, table_range in sheet.tables.items():
                try:
                    table_data = self._extract_table_range(sheet, table_range, sheet_name, table_name)
                    if table_data:
                        tables.append(table_data)
                except Exception as e:
                    logger.warning(f"Failed to extract table {table_name} from sheet {sheet_name}: {e}")
        
        # If no defined tables, treat entire sheet as a table
        if not tables and sheet.max_row and sheet.max_column:
            try:
                # Create a range string for the entire data area
                range_string = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                table_data = self._extract_table_range(sheet, range_string, sheet_name, "full_sheet")
                if table_data:
                    tables.append(table_data)
            except Exception as e:
                logger.warning(f"Failed to extract full sheet as table from {sheet_name}: {e}")
        
        return tables
    
    def _extract_table_range(
        self,
        sheet: Worksheet,
        table_range: str,
        sheet_name: str,
        table_name: str
    ) -> Optional[Dict[str, Any]]:
        """Extract data from a specific table range.
        
        Args:
            sheet: Worksheet containing the table
            table_range: Range string (e.g., "A1:D10")
            sheet_name: Name of the worksheet
            table_name: Name of the table
            
        Returns:
            Dictionary with table data or None if extraction fails
        """
        try:
            # Parse the range
            cells = list(sheet[table_range])
            
            if not cells:
                return None
            
            # Apply row/column limits
            max_rows = self.max_rows_per_table if self.max_rows_per_table > 0 else len(cells)
            max_cols = self.max_columns_per_table if self.max_columns_per_table > 0 else len(cells[0]) if cells else 0
            
            # Extract table data
            table_data = {
                "sheet_name": sheet_name,
                "table_name": table_name,
                "range": table_range,
                "rows": min(len(cells), max_rows),
                "columns": min(len(cells[0]) if cells else 0, max_cols),
                "data": []
            }
            
            # Extract cell values
            for i, row in enumerate(cells):
                if i >= max_rows:
                    table_data["truncated_rows"] = True
                    break
                    
                row_data = []
                for j, cell in enumerate(row):
                    if j >= max_cols:
                        table_data["truncated_columns"] = True
                        break
                    row_data.append(cell.value)
                table_data["data"].append(row_data)
            
            return table_data
            
        except Exception as e:
            logger.warning(f"Failed to extract table range {table_range} from sheet {sheet_name}: {e}")
            return None
    
    def _extract_formulas_from_sheet(self, sheet: Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract formulas from a worksheet.
        
        Args:
            sheet: Worksheet to extract formulas from
            sheet_name: Name of the worksheet
            
        Returns:
            List of formula dictionaries
        """
        formulas = []
        
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            "sheet_name": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "row": cell.row,
                            "column": cell.column
                        })
        except Exception as e:
            logger.warning(f"Failed to extract formulas from sheet {sheet_name}: {e}")
        
        return formulas
    
    def _extract_comments_from_sheet(self, sheet: Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract comments from a worksheet.
        
        Args:
            sheet: Worksheet to extract comments from
            sheet_name: Name of the worksheet
            
        Returns:
            List of comment dictionaries
        """
        comments = []
        
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        comments.append({
                            "sheet_name": sheet_name,
                            "cell": cell.coordinate,
                            "author": cell.comment.author if hasattr(cell.comment, 'author') else None,
                            "text": cell.comment.text,
                            "row": cell.row,
                            "column": cell.column
                        })
        except Exception as e:
            logger.warning(f"Failed to extract comments from sheet {sheet_name}: {e}")
        
        return comments
    
    def _extract_images_from_sheet(
        self,
        sheet: Worksheet,
        sheet_name: str,
        start_index: int
    ) -> List[Dict[str, Any]]:
        """Extract images from a worksheet.
        
        Args:
            sheet: Worksheet to extract images from
            sheet_name: Name of the worksheet
            start_index: Starting index for image numbering
            
        Returns:
            List of image dictionaries
        """
        images = []
        
        try:
            if hasattr(sheet, '_images') and sheet._images:
                for idx, image in enumerate(sheet._images):
                    try:
                        # Get image data
                        image_bytes = image._data()
                        
                        # Prepare image data based on output mode
                        if self.image_output_mode == "base64":
                            image_data = base64.b64encode(image_bytes).decode('utf-8')
                        else:
                            image_data = image_bytes
                        
                        image_info = {
                            "sheet_name": sheet_name,
                            "image_index": start_index + idx,
                            "format": image.format if hasattr(image, 'format') else 'unknown',
                            "size_bytes": len(image_bytes),
                            "data": image_data,
                            "encoding": self.image_output_mode
                        }
                        
                        # Add anchor information if available
                        if hasattr(image, 'anchor'):
                            anchor = image.anchor
                            if hasattr(anchor, '_from'):
                                image_info["anchor_cell"] = f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"
                        
                        images.append(image_info)
                        
                    except Exception as e:
                        logger.warning(f"Failed to extract image {idx} from sheet {sheet_name}: {e}")
        except Exception as e:
            logger.warning(f"Failed to access images from sheet {sheet_name}: {e}")
        
        return images
    
    def _extract_workbook_properties(self, wb) -> Dict[str, Any]:
        """Extract workbook properties and metadata.
        
        Args:
            wb: Opened openpyxl Workbook
            
        Returns:
            Dictionary of workbook properties
        """
        properties = {}
        
        try:
            if hasattr(wb, 'properties'):
                props = wb.properties
                
                properties['title'] = props.title or None
                properties['creator'] = props.creator or None
                properties['subject'] = props.subject or None
                properties['keywords'] = props.keywords or None
                properties['description'] = props.description or None
                properties['category'] = props.category or None
                properties['created'] = props.created.isoformat() if props.created else None
                properties['modified'] = props.modified.isoformat() if props.modified else None
                properties['last_modified_by'] = props.lastModifiedBy or None
                properties['revision'] = props.revision
                properties['version'] = props.version or None
            
            # Add workbook-specific info
            properties['sheet_count'] = len(wb.sheetnames)
            properties['sheet_names'] = wb.sheetnames
            
            # Add named ranges if available
            if hasattr(wb, 'defined_names') and wb.defined_names:
                properties['named_ranges'] = [name for name in wb.defined_names.definedName]
            
        except Exception as e:
            logger.warning(f"Failed to extract some workbook properties: {e}")
        
        return properties
